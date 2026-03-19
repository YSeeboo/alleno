from io import BytesIO
from urllib.parse import quote
from zipfile import ZipFile

from openpyxl import load_workbook
from PIL import Image

from models.plating_order import PlatingOrder
from services.inventory import add_stock
from services.part import create_part


def test_download_plating_excel_exports_template_content(client, db, monkeypatch):
    def fake_download_image_bytes(source):
        image = Image.new("RGB", (8, 8), _color_for_source(source))
        buffer = BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()

    monkeypatch.setattr("services.plating_excel.download_image_bytes", fake_download_image_bytes)

    part = create_part(
        db,
        {
            "name": "蝴蝶扣",
            "category": "小配件",
            "image": "https://img.example.com/part.png",
            "unit": "个",
        },
    )
    add_stock(db, "part", part.id, 20.0, "initial stock")
    db.flush()

    create_resp = client.post(
        "/api/plating/",
        json={
            "supplier_name": "广发电镀厂",
            "items": [
                {
                    "part_id": part.id,
                    "qty": 12,
                    "unit": "个",
                    "plating_method": "白K",
                    "note": "挂镀，避免刮花",
                }
            ],
            "note": "订单级备注不会写到明细要求",
        },
    )
    order_id = create_resp.json()["id"]

    patch_resp = client.patch(
        f"/api/plating/{order_id}/delivery-images",
        json={
            "delivery_images": [
                "https://img.example.com/delivery-1.png",
                "https://img.example.com/delivery-2.png",
            ]
        },
    )
    assert patch_resp.status_code == 200

    response = client.get(f"/api/plating/{order_id}/excel")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="plating-export.xlsx"' in response.headers["content-disposition"]
    assert 'filename*=UTF-8' in response.headers["content-disposition"]
    assert '.xlsx' in response.headers["content-disposition"]

    media_files = [
        name
        for name in ZipFile(BytesIO(response.content)).namelist()
        if name.startswith("xl/media/") and name.lower().endswith(".png")
    ]
    assert len(media_files) == 4

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    order = db.query(PlatingOrder).filter(PlatingOrder.id == order_id).first()

    assert sheet["A6"].value == "顾客: 广发电镀厂"
    assert sheet["C6"].value == f"{order.created_at.year:04d}年 {order.created_at.month:02d}月 {order.created_at.day:02d}日"
    assert sheet["A8"].value == "蝴蝶扣"
    assert sheet["C8"].value == "白K"
    assert float(sheet["D8"].value) == 12.0
    assert sheet["E8"].value == "个"
    assert sheet["F8"].value == "挂镀，避免刮花"
    assert float(sheet["E8"].font.sz) == 18.0
    assert float(sheet["F8"].font.sz) == 18.0
    assert sheet["A13"].value == "发货图片："


def test_download_plating_excel_uses_supplier_and_short_date_in_filename(client, db):
    part = create_part(
        db,
        {
            "name": "文件名配件",
            "category": "小配件",
            "unit": "个",
        },
    )
    add_stock(db, "part", part.id, 10.0, "initial stock")
    db.flush()

    create_resp = client.post(
        "/api/plating/",
        json={
            "supplier_name": 'A/B:电镀厂',
            "items": [{"part_id": part.id, "qty": 2, "unit": "个"}],
        },
    )
    order_id = create_resp.json()["id"]
    order = db.query(PlatingOrder).filter(PlatingOrder.id == order_id).first()

    response = client.get(f"/api/plating/{order_id}/excel")
    assert response.status_code == 200

    expected_date = f"{order.created_at.year % 100:02d}{order.created_at.month:02d}{order.created_at.day:02d}"
    expected_filename = f"发出_A_B_电镀厂_{expected_date}.xlsx"
    assert response.headers["content-disposition"] == (
        f'attachment; filename="plating-export.xlsx"; filename*=UTF-8\'\'{quote(expected_filename)}'
    )


def test_download_plating_excel_expands_detail_rows(client, db):
    items = []
    for index in range(6):
        part = create_part(
            db,
            {
                "name": f"配件{index + 1}",
                "category": "小配件",
                "unit": "个",
            },
        )
        add_stock(db, "part", part.id, 20.0, "initial stock")
        items.append(
            {
                "part_id": part.id,
                "qty": index + 1,
                "unit": "个",
                "plating_method": f"颜色{index + 1}",
                "note": f"备注{index + 1}",
            }
        )
    db.flush()

    create_resp = client.post(
        "/api/plating/",
        json={
            "supplier_name": "扩展测试电镀厂",
            "items": items,
        },
    )
    order_id = create_resp.json()["id"]

    response = client.get(f"/api/plating/{order_id}/excel")
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active

    assert sheet["A8"].value == "配件1"
    assert sheet["A13"].value == "配件6"
    assert float(sheet["E13"].font.sz) == 18.0
    assert float(sheet["F13"].font.sz) == 18.0
    assert sheet["A14"].value == "发货图片："
    assert sheet.max_row == 57
    assert "A14:F17" in {str(item) for item in sheet.merged_cells.ranges}
    assert "A18:D37" in {str(item) for item in sheet.merged_cells.ranges}
    assert "E18:F37" in {str(item) for item in sheet.merged_cells.ranges}


def test_download_plating_excel_order_not_found(client):
    response = client.get("/api/plating/EP-9999/excel")
    assert response.status_code == 404


def _color_for_source(source: str) -> tuple[int, int, int]:
    total = sum(ord(char) for char in source)
    return (
        40 + total % 150,
        40 + (total * 3) % 150,
        40 + (total * 5) % 150,
    )
