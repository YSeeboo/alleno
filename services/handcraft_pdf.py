from __future__ import annotations

import math
from functools import lru_cache
from io import BytesIO

from openpyxl import load_workbook
from PIL import Image as PILImage, UnidentifiedImageError
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader, simpleSplit
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.pdfmetrics import registerFont, stringWidth
from reportlab.pdfgen import canvas

from services.handcraft_export import (
    build_export_filename,
    download_pdf_image_bytes as download_image_bytes,
    format_excel_date,
    get_handcraft_export_payload,
)
from services.plating_excel import _TEMPLATE_PATH

_PAGE_WIDTH, _PAGE_HEIGHT = A4
_MARGIN_X = 48
_MARGIN_TOP = 32
_MARGIN_BOTTOM = 30
_CONTENT_WIDTH = _PAGE_WIDTH - _MARGIN_X * 2
_FIRST_PAGE_ROW_HEIGHT = 32
_DETAIL_PAGE_ROW_HEIGHT = 32
_DETAIL_PAGE_MAX_ROWS = 23
_HEADER_ROW_HEIGHT = 28
_IMAGE_PADDING = 2
_MAX_IMAGE_HEIGHT = 200
_MIN_IMAGE_HEIGHT = 80
_IMAGE_GAP = 10
_IMAGE_ROW_GAP = 28
_LABEL_FONT = "STSong-Light"
_LABEL_FONT_BOLD = "Helvetica-Bold"


def _compute_image_layout(image_count: int, available_space: float):
    """Return (cols, rows, image_height) or None if images don't fit.

    - 1 image: single column
    - 2-4 images: 2 columns, ceil(n/2) rows
    - 5+ images: single column, fit as many as possible (may need multi-page)
    """
    if image_count == 0:
        return None

    space = available_space - _IMAGE_GAP
    if space <= 0:
        return None

    if image_count == 1:
        h = min(space, _MAX_IMAGE_HEIGHT)
        return (1, 1, h) if h >= _MIN_IMAGE_HEIGHT else None

    if image_count <= 4:
        # 2-4: try 2 columns
        if image_count == 2:
            h_single = (space - _IMAGE_ROW_GAP) / 2
            if h_single >= _MAX_IMAGE_HEIGHT:
                return (1, 2, _MAX_IMAGE_HEIGHT)
            h_double = min(space, _MAX_IMAGE_HEIGHT)
            return (2, 1, h_double) if h_double >= _MIN_IMAGE_HEIGHT else None
        import math
        rows = math.ceil(image_count / 2)
        h = (space - _IMAGE_ROW_GAP * (rows - 1)) / rows
        h = min(h, _MAX_IMAGE_HEIGHT)
        return (2, rows, h) if h >= _MIN_IMAGE_HEIGHT else None

    # 5+ images: single column, max 3 per page, overflow to next page
    rows = min(image_count, 3)
    h = (space - _IMAGE_ROW_GAP * (rows - 1)) / rows
    h = min(h, _MAX_IMAGE_HEIGHT)
    return (1, rows, h) if h >= _MIN_IMAGE_HEIGHT else None


def _max_rows_with_images(image_count: int, total_available: float, row_height: float) -> int:
    """Max data rows that fit while also fitting images."""
    if image_count == 0:
        return int(total_available // row_height)

    max_possible = int(total_available // row_height)
    for rows in range(max_possible, 0, -1):
        remaining = total_available - rows * row_height
        layout = _compute_image_layout(image_count, remaining)
        if layout is not None:
            return rows
    return 0


def build_handcraft_order_pdf(db, order_id: str) -> tuple[bytes, str]:
    payload = get_handcraft_export_payload(db, order_id)
    template_text = _load_template_text()
    _register_fonts()

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    pdf.setTitle(build_export_filename(payload["order"].supplier_name, payload["order"].created_at, "pdf"))

    details = payload["details"]
    for i, d in enumerate(details, 1):
        d["seq"] = i
    delivery_images = payload["delivery_images"]

    # Render order is fixed: parts → shortage section → images. Inline-image
    # optimization (squeezing ≤4 images onto the same page as parts) was
    # dropped because it forces images to render BEFORE the shortage section,
    # which contradicts the intended ordering. ≤4 images can still flow
    # inline below the shortage section if there's room on the same page.

    first_page_available = _measure_first_page_available(pdf, payload, template_text)
    detail_page_available = _PAGE_HEIGHT - _MARGIN_TOP - _MARGIN_BOTTOM - _HEADER_ROW_HEIGHT

    # Phase 1: parts pages (rows only, no images on parts pages anymore).
    last_y = _draw_parts_pages(
        pdf, payload, template_text, details,
        first_page_available=first_page_available,
        detail_page_available=detail_page_available,
    )

    # Phase 2: shortage section / "全部配齐" notice.
    last_y = _draw_shortage_section(pdf, payload, last_y)

    # Phase 3: delivery images. ≤4 images may flow inline below the shortage
    # section; 5+ always get dedicated pages.
    if delivery_images:
        last_y = _draw_delivery_images(pdf, payload, template_text, delivery_images, last_y)

    # Phase 4: 手工回执 page — appears only when there are jewelry breakdown
    # entries with a resolved customer. Pure metadata, supplier-facing.
    _draw_handcraft_receipt_page(pdf, db, payload["order"])

    pdf.save()
    return buffer.getvalue(), build_export_filename(payload["order"].supplier_name, payload["order"].created_at, "pdf")


def _draw_parts_pages(
    pdf, payload: dict, template_text: dict, details: list[dict],
    first_page_available: float, detail_page_available: float,
) -> float:
    """Paginate the parts table across pages without any inline images.
    Returns the final y position after the last drawn row."""
    if not details:
        # Empty order — still render a header so the page isn't blank.
        return _draw_detail_page(
            pdf, payload, template_text, [],
            include_static_header=True, page_images=[],
        )

    remaining = list(details)
    page_index = 0
    last_y = _PAGE_HEIGHT - _MARGIN_TOP

    while remaining:
        if page_index == 0:
            row_h = _FIRST_PAGE_ROW_HEIGHT
            page_size = _max_rows_with_images(0, first_page_available, row_h)
        else:
            row_h = _DETAIL_PAGE_ROW_HEIGHT
            page_size = _DETAIL_PAGE_MAX_ROWS

        chunk = remaining[:page_size]
        remaining = remaining[page_size:]

        last_y = _draw_detail_page(
            pdf, payload, template_text, chunk,
            include_static_header=(page_index == 0),
            page_images=[],
        )

        if remaining:
            pdf.showPage()
            last_y = _PAGE_HEIGHT - _MARGIN_TOP
        page_index += 1

    return last_y


def _draw_delivery_images(
    pdf, payload: dict, template_text: dict,
    delivery_images: list[str], current_y: float,
) -> float:
    """Render delivery images. ≤4 images: try to fit inline below current_y;
    if there's no room, dedicated page(s). 5+ images: always dedicated pages."""
    image_count = len(delivery_images)
    space_left = current_y - _MARGIN_BOTTOM

    # ≤4 images may flow inline below the shortage section if there's room.
    if image_count <= 4:
        layout = _compute_image_layout(image_count, space_left)
        if layout is not None:
            cols, rows, image_height = layout
            current_y -= _IMAGE_GAP
            _draw_images_block(pdf, delivery_images, current_y, cols, rows, image_height)
            return current_y - rows * image_height - (rows - 1) * _IMAGE_ROW_GAP

    # Either too many images to fit inline, or 5+ — go to dedicated page(s).
    pdf.showPage()
    _draw_images_page(pdf, payload, template_text, delivery_images)
    return _MARGIN_BOTTOM


# ---------------------------------------------------------------------------
# Shortage section (缺件清单 / 全部配齐文案)
# ---------------------------------------------------------------------------


_SHORTAGE_TITLE = "本单缺件清单"
_SHORTAGE_SUBTITLE = "以下配件因库存不足暂未发出，待补货完成后会另行发送。"
_NO_SHORTAGE_TEXT = "配件已经全部配齐，收到配件后请清点配件"
_SHORTAGE_TITLE_HEIGHT = 26
_SHORTAGE_SUBTITLE_HEIGHT = 18
_SHORTAGE_ROW_HEIGHT = 32
_SHORTAGE_TABLE_HEADER_HEIGHT = 28
_NO_SHORTAGE_BLOCK_HEIGHT = 80


def _draw_shortage_section(pdf, payload: dict, current_y: float) -> float:
    """Draw the shortage section (or the all-clear notice) starting at current_y.
    Flows inline if room, else starts a new page. Multi-page table chunks
    continue without re-rendering the title/header (per design)."""
    rows = payload.get("shortage_rows") or []
    if not rows:
        return _draw_no_shortage_notice(pdf, current_y)
    return _draw_shortage_table(pdf, rows, current_y)


def _draw_no_shortage_notice(pdf, current_y: float) -> float:
    """Center-aligned star icon + bold 18px text. Always fits on a page."""
    if current_y - _NO_SHORTAGE_BLOCK_HEIGHT < _MARGIN_BOTTOM:
        pdf.showPage()
        current_y = _PAGE_HEIGHT - _MARGIN_TOP

    block_top = current_y - 16  # leading gap from previous content
    block_bottom = block_top - _NO_SHORTAGE_BLOCK_HEIGHT
    cy = (block_top + block_bottom) / 2

    icon_size = 42
    gap = 14
    pdf.setFont(_LABEL_FONT, 18)
    text_width = stringWidth(_NO_SHORTAGE_TEXT, _LABEL_FONT, 18)
    total_width = icon_size + gap + text_width
    start_x = (_PAGE_WIDTH - total_width) / 2

    icon_y = cy - icon_size / 2
    _draw_star_icon(pdf, start_x, icon_y, icon_size)

    pdf.setFillColor(colors.black)
    text_y = cy - 6  # baseline tweak so visual center matches icon
    pdf.drawString(start_x + icon_size + gap, text_y, _NO_SHORTAGE_TEXT)

    return block_bottom


def _draw_shortage_table(pdf, rows: list[dict], current_y: float) -> float:
    """Render the shortage table. Title + subtitle on the first page only;
    subsequent pages just continue the body rows (per design)."""
    widths = _shortage_column_widths()
    headers = ["序号", "名称", "图片", "颜色", "数量", "备注"]
    leading_gap = 12
    title_block_height = leading_gap + _SHORTAGE_TITLE_HEIGHT + _SHORTAGE_SUBTITLE_HEIGHT + _SHORTAGE_TABLE_HEADER_HEIGHT

    # If the title block alone won't fit, push to a new page first.
    if current_y - title_block_height - _SHORTAGE_ROW_HEIGHT < _MARGIN_BOTTOM:
        pdf.showPage()
        current_y = _PAGE_HEIGHT - _MARGIN_TOP

    # Title
    y = current_y - leading_gap
    pdf.setFont(_LABEL_FONT, 14)
    pdf.setFillColor(colors.black)
    title_width = stringWidth(_SHORTAGE_TITLE, _LABEL_FONT, 14)
    pdf.drawString((_PAGE_WIDTH - title_width) / 2, y - 16, _SHORTAGE_TITLE)
    y -= _SHORTAGE_TITLE_HEIGHT

    # Subtitle
    pdf.setFont(_LABEL_FONT, 9)
    pdf.setFillColor(colors.HexColor("#666666"))
    pdf.drawString(_MARGIN_X, y - 12, _SHORTAGE_SUBTITLE)
    y -= _SHORTAGE_SUBTITLE_HEIGHT

    # Table header
    _draw_shortage_table_header(pdf, headers, y, widths)
    y -= _SHORTAGE_TABLE_HEADER_HEIGHT

    # Body rows; paginate without re-rendering header per user's choice
    for row in rows:
        if y - _SHORTAGE_ROW_HEIGHT < _MARGIN_BOTTOM:
            pdf.showPage()
            y = _PAGE_HEIGHT - _MARGIN_TOP
        _draw_shortage_row(pdf, row, y, widths, _SHORTAGE_ROW_HEIGHT)
        y -= _SHORTAGE_ROW_HEIGHT

    return y


def _shortage_column_widths() -> list[float]:
    # 序号 / 名称 / 图片 / 颜色 / 数量 / 备注
    raw = [7.0, 26.0, 16.0, 12.0, 12.0, 27.0]
    total = sum(raw)
    scaled = [_CONTENT_WIDTH * v / total for v in raw]
    scaled[-1] = _CONTENT_WIDTH - sum(scaled[:-1])
    return scaled


def _draw_shortage_table_header(pdf, headers: list[str], top_y: float, widths: list[float]) -> None:
    x = _MARGIN_X
    pdf.setStrokeColor(colors.black)
    pdf.setLineWidth(0.5)
    for header, width in zip(headers, widths):
        pdf.setFillColor(colors.HexColor("#d9d9d9"))
        pdf.rect(x, top_y - _SHORTAGE_TABLE_HEADER_HEIGHT, width, _SHORTAGE_TABLE_HEADER_HEIGHT, stroke=1, fill=1)
        pdf.setFillColor(colors.black)
        pdf.setFont(_LABEL_FONT, 10)
        text_width = stringWidth(header, _LABEL_FONT, 10)
        pdf.drawString(
            x + max(0, (width - text_width) / 2),
            top_y - _SHORTAGE_TABLE_HEADER_HEIGHT + 9,
            header,
        )
        x += width


def _draw_shortage_row(pdf, row: dict, top_y: float, widths: list[float], row_height: float) -> None:
    x = _MARGIN_X
    pdf.setStrokeColor(colors.black)
    pdf.setLineWidth(0.5)
    pdf.setFillColor(colors.white)
    for w in widths:
        pdf.rect(x, top_y - row_height, w, row_height, stroke=1, fill=0)
        x += w

    cells = _cell_positions(widths, top_y, row_height)
    pdf.setFillColor(colors.black)
    _draw_centered_text(pdf, str(row.get("seq", "")), *cells[0], font_size=10)
    name = row.get("name") or row.get("part_id") or ""
    _draw_centered_paragraph(pdf, name, *cells[1], font_size=10, max_lines=2)
    if row.get("part_image"):
        _draw_image_in_box(pdf, row["part_image"], *cells[2])
    _draw_centered_text(pdf, row.get("color") or "", *cells[3], font_size=10)
    _draw_centered_text(pdf, _format_shortage_qty(row.get("qty")), *cells[4], font_size=10)
    _draw_centered_paragraph(pdf, row.get("note") or "", *cells[5], font_size=10, max_lines=2)


def _format_shortage_qty(v) -> str:
    if v is None:
        return "—"
    f = float(v)
    if f == int(f):
        return str(int(f))
    return f"{f:.4f}".rstrip("0").rstrip(".")


def _draw_star_icon(pdf, x: float, y: float, size: float) -> None:
    """Draw a 5-point yellow star inside an `size`-by-`size` box anchored at (x, y) bottom-left."""
    cx = x + size / 2
    cy = y + size / 2
    r_outer = size / 2
    # Inner radius ratio that matches a "classic" sharp star appearance.
    r_inner = r_outer * 0.382

    pdf.saveState()
    pdf.setFillColor(colors.HexColor("#F3D958"))
    pdf.setStrokeColor(colors.HexColor("#F3D958"))
    pdf.setLineWidth(0.3)

    path = pdf.beginPath()
    for i in range(10):
        angle = -math.pi / 2 + i * math.pi / 5  # start at the top, alternate outer/inner
        r = r_outer if i % 2 == 0 else r_inner
        px = cx + r * math.cos(angle)
        py = cy + r * math.sin(angle)
        if i == 0:
            path.moveTo(px, py)
        else:
            path.lineTo(px, py)
    path.close()
    pdf.drawPath(path, stroke=1, fill=1)
    pdf.restoreState()


def _measure_first_page_available(pdf, payload: dict, template_text: dict) -> float:
    """Calculate available space on first page after static header + table header."""
    y = _PAGE_HEIGHT - _MARGIN_TOP

    if template_text["logo_bytes"]:
        logo = _fit_image(template_text["logo_bytes"], 116, 38)
        if logo is not None:
            _, _, draw_height = logo
            y -= draw_height + 12

    for line in template_text["header_lines"]:
        if line:
            y -= 14

    y -= 10  # gap before supplier line
    y -= 10  # supplier line to table

    y -= _HEADER_ROW_HEIGHT
    return y - _MARGIN_BOTTOM


def _draw_detail_page(pdf, payload: dict, template_text: dict, details: list[dict],
                      include_static_header: bool, page_images: list[str]) -> float:
    """Draw the parts table page. Returns final y position (after rows + any
    inline images), so callers can place the shortage section right below."""
    y = _PAGE_HEIGHT - _MARGIN_TOP

    if include_static_header:
        y = _draw_static_header(pdf, payload, template_text, y)
    # No compact header on continuation pages — start directly with table

    table_top = y
    column_widths = _column_widths()
    _draw_table_header(pdf, template_text["detail_headers"], table_top, column_widths)
    y = table_top - _HEADER_ROW_HEIGHT

    row_height = _FIRST_PAGE_ROW_HEIGHT if include_static_header else _DETAIL_PAGE_ROW_HEIGHT
    for detail in details:
        _draw_detail_row(pdf, detail, y, column_widths, row_height)
        y -= row_height

    if page_images:
        image_count = len(page_images)
        remaining_space = y - _MARGIN_BOTTOM
        layout = _compute_image_layout(image_count, remaining_space)
        if layout is not None:
            cols, rows, image_height = layout
            y -= _IMAGE_GAP
            _draw_images_block(pdf, page_images, y, cols, rows, image_height)
            # The image block consumes (rows * height + (rows-1) * gap)
            y -= rows * image_height + (rows - 1) * _IMAGE_ROW_GAP

    return y


def _draw_images_page(pdf, payload: dict, template_text: dict, delivery_images: list[str]) -> None:
    remaining = list(delivery_images)
    while remaining:
        y = _PAGE_HEIGHT - _MARGIN_TOP
        available = y - _MARGIN_BOTTOM
        layout = _compute_image_layout(len(remaining), available)
        if layout is None:
            break
        cols, rows, image_height = layout
        drawn = cols * rows
        _draw_images_block(pdf, remaining[:drawn], y, cols, rows, image_height)
        remaining = remaining[drawn:]
        if remaining:
            pdf.showPage()


def _draw_images_block(pdf, images: list[str], y: float, cols: int, rows: int, image_height: float) -> None:
    """Draw images in a grid layout starting from y downward."""
    gutter = 26
    if cols == 1:
        box_width = _CONTENT_WIDTH
    else:
        box_width = (_CONTENT_WIDTH - gutter) / 2

    for index, source in enumerate(images[:cols * rows]):
        col = index % cols
        row = index // cols
        x = _MARGIN_X + col * (box_width + gutter)
        top_y = y - row * (image_height + _IMAGE_ROW_GAP)
        box_y = top_y - image_height
        _draw_image_in_box(pdf, source, x, box_y, box_width, image_height)


def _draw_static_header(pdf, payload: dict, template_text: dict, y: float) -> float:
    if template_text["logo_bytes"]:
        logo = _fit_image(template_text["logo_bytes"], 116, 38)
        if logo is not None:
            image_reader, draw_width, draw_height = logo
            pdf.drawImage(
                image_reader,
                (_PAGE_WIDTH - draw_width) / 2,
                y - draw_height,
                width=draw_width,
                height=draw_height,
                preserveAspectRatio=True,
                mask="auto",
            )
            y -= draw_height + 12

    pdf.setFont(_LABEL_FONT, 10)
    pdf.setFillColor(colors.black)
    for line in template_text["header_lines"]:
        if line:
            pdf.drawString(_MARGIN_X, y, line)
            y -= 14

    y -= 10
    supplier_text = f"{template_text['customer_prefix']}{payload['order'].supplier_name or ''}"
    date_text = format_excel_date(payload["order"].created_at) or ""
    pdf.setFont(_LABEL_FONT, 10)
    pdf.drawString(_MARGIN_X, y, supplier_text)
    text_width = stringWidth(date_text, _LABEL_FONT, 10)
    pdf.drawString(_PAGE_WIDTH - _MARGIN_X - text_width, y, date_text)
    return y - 10


def _draw_table_header(pdf, headers: list[str], top_y: float, column_widths: list[float]) -> None:
    x = _MARGIN_X
    pdf.setStrokeColor(colors.HexColor("#000000"))
    pdf.setLineWidth(0.5)
    for index, header in enumerate(headers):
        width = column_widths[index]
        pdf.setFillColor(colors.HexColor("#d9d9d9"))
        pdf.rect(x, top_y - _HEADER_ROW_HEIGHT, width, _HEADER_ROW_HEIGHT, stroke=1, fill=1)
        _draw_header_text(pdf, header, x + 4, top_y - _HEADER_ROW_HEIGHT + 3, width - 8, _HEADER_ROW_HEIGHT - 6)
        x += width


def _draw_detail_row(pdf, detail: dict, top_y: float, column_widths: list[float], row_height: float) -> None:
    x = _MARGIN_X
    pdf.setStrokeColor(colors.HexColor("#000000"))
    pdf.setLineWidth(0.5)
    pdf.setFillColor(colors.white)
    for width in column_widths:
        pdf.rect(x, top_y - row_height, width, row_height, stroke=1, fill=0)
        x += width

    cells = _cell_positions(column_widths, top_y, row_height)
    pdf.setFillColor(colors.black)
    _draw_centered_text(pdf, str(detail.get("seq", "")), *cells[0], font_size=10)
    _draw_centered_paragraph(pdf, detail["name"] or detail["part_id"], *cells[1], font_size=10, max_lines=2)
    _draw_image_in_box(pdf, detail["part_image"], *cells[2])
    _draw_centered_text(pdf, detail["plating_method"], *cells[3], font_size=10)
    _draw_centered_text(pdf, detail["qty_text"], *cells[4], font_size=10)
    _draw_centered_text(pdf, detail["unit"], *cells[5], font_size=10)
    _draw_centered_text(pdf, detail.get("weight_text", ""), *cells[6], font_size=10)
    _draw_centered_paragraph(pdf, detail["note"], *cells[7], font_size=10, max_lines=2)


def _draw_image_in_box(pdf, source: str, x: float, y: float, width: float, height: float) -> None:
    image_bytes = download_image_bytes(source)
    if not image_bytes:
        return
    placement = _fit_image(image_bytes, width - _IMAGE_PADDING * 2, height - _IMAGE_PADDING * 2)
    if placement is None:
        return

    image_reader, draw_width, draw_height = placement
    offset_x = x + (width - draw_width) / 2
    offset_y = y + (height - draw_height) / 2
    pdf.drawImage(
        image_reader,
        offset_x,
        offset_y,
        width=draw_width,
        height=draw_height,
        preserveAspectRatio=True,
        mask="auto",
    )


def _fit_image(image_bytes: bytes, max_width: float, max_height: float):
    try:
        with PILImage.open(BytesIO(image_bytes)) as raw_image:
            image = raw_image.copy()
    except (UnidentifiedImageError, OSError):
        return None

    img_width, img_height = image.size
    if img_width <= 0 or img_height <= 0:
        return None

    scale = min(max_width / img_width, max_height / img_height)
    draw_width = max(1, img_width * scale)
    draw_height = max(1, img_height * scale)
    return ImageReader(BytesIO(image_bytes)), draw_width, draw_height


def _draw_centered_text(pdf, text: str, x: float, y: float, width: float, height: float, font_size: int) -> None:
    lines = simpleSplit(text or "", _LABEL_FONT, font_size, max(width, 1))[:2]
    if not lines:
        return
    pdf.setFont(_LABEL_FONT, font_size)
    total_height = len(lines) * (font_size + 2)
    current_y = y + (height + total_height) / 2 - font_size
    for line in lines:
        line_width = stringWidth(line, _LABEL_FONT, font_size)
        pdf.drawString(x + max(0, (width - line_width) / 2), current_y, line)
        current_y -= font_size + 2


def _draw_centered_paragraph(pdf, text: str, x: float, y: float, width: float, height: float, font_size: int, max_lines: int) -> None:
    lines = simpleSplit(text or "", _LABEL_FONT, font_size, max(width - 6, 1))[:max_lines]
    if not lines:
        return
    pdf.setFont(_LABEL_FONT, font_size)
    line_gap = font_size + 2
    total_height = len(lines) * line_gap
    current_y = y + (height + total_height) / 2 - font_size
    for line in lines:
        line_width = stringWidth(line, _LABEL_FONT, font_size)
        pdf.drawString(x + max(0, (width - line_width) / 2), current_y, line)
        current_y -= line_gap


def _cell_positions(column_widths: list[float], top_y: float, row_height: float) -> list[tuple[float, float, float, float]]:
    positions = []
    x = _MARGIN_X
    for width in column_widths:
        positions.append((x, top_y - row_height, width, row_height))
        x += width
    return positions


def _column_widths() -> list[float]:
    # 序号, 名称, 图片, 颜色, 数量, 单位, 重量, 备注
    raw_widths = [6.0, 32.0, 18.6923, 18.6923, 13.0, 10.0, 16.0, 50.0]
    total = sum(raw_widths)
    scaled = [(_CONTENT_WIDTH * value / total) for value in raw_widths]
    scaled[-1] = _CONTENT_WIDTH - sum(scaled[:-1])
    return scaled


@lru_cache(maxsize=1)
def _load_template_text() -> dict:
    workbook = load_workbook(_TEMPLATE_PATH, data_only=True)
    sheet = workbook.active
    header_lines = [sheet["A2"].value or "", sheet["A3"].value or "", sheet["A4"].value or ""]
    raw_headers = ["序号"] + [sheet.cell(row=7, column=column).value or "" for column in range(1, 7)]
    # Insert "重量" column after 单位 (index 5), before 备注 (index 6)
    detail_headers = raw_headers[:6] + ["重量"] + raw_headers[6:]
    customer_prefix = sheet["A6"].value or "顾客: "
    delivery_title = sheet["A13"].value or "发货图片："
    logo_bytes = None
    images = getattr(sheet, "_images", [])
    if images:
        try:
            logo_bytes = images[0]._data()
        except Exception:
            logo_bytes = None
    workbook.close()
    return {
        "header_lines": header_lines,
        "detail_headers": detail_headers,
        "customer_prefix": customer_prefix,
        "delivery_title": delivery_title,
        "logo_bytes": logo_bytes,
    }


@lru_cache(maxsize=1)
def _register_fonts() -> bool:
    registerFont(UnicodeCIDFont(_LABEL_FONT))
    return True


def _draw_header_text(pdf, text: str, x: float, y: float, width: float, height: float) -> None:
    parts = (text or "").split("\n", 1)
    title = parts[0] if parts else ""
    subtitle = parts[1] if len(parts) > 1 else ""

    if title:
        pdf.setFont(_LABEL_FONT, 8)
        pdf.setFillColor(colors.black)
        title_width = stringWidth(title, _LABEL_FONT, 8)
        pdf.drawString(x + max(0, (width - title_width) / 2), y + height - 13, title)

    if subtitle:
        pdf.setFont("Helvetica", 4.5)
        pdf.setFillColor(colors.black)
        subtitle_width = stringWidth(subtitle, "Helvetica", 4.5)
        pdf.drawString(x + max(0, (width - subtitle_width) / 2), y + 4, subtitle)


# ──────────────────────────────────────────────────────────────────────
# 手工回执 page (customer breakdown for sorting returned goods).
# Real customer names are intentionally replaced with per-HC sequential
# aliases "客户 N" so the supplier cannot correlate customer identities
# across multiple HCs over time.
# ──────────────────────────────────────────────────────────────────────

_RECEIPT_TITLE = "手工回执"
_RECEIPT_NOTE = "请于回货时随成品一并交回此单以便分拣核对"

# Receipt page typography & spacing
_RP_INK = colors.HexColor("#111111")
_RP_INK_SOFT = colors.HexColor("#555555")
_RP_RULE = colors.HexColor("#666666")
_RP_RULE_SOFT = colors.HexColor("#d8d8d8")
_RP_FOOTER_BG = colors.HexColor("#f3f3f3")
_RP_FOOTER_RULE = colors.HexColor("#c4c4c4")
_RP_ROW_H = 38                  # customer row height in the breakdown table
_RP_HEADER_H = 26               # header row height in the breakdown table
# Column proportions of _CONTENT_WIDTH for the 5-column table
_RP_COL_PCT = (0.26, 0.12, 0.18, 0.24, 0.20)


def _wrap_text_to_lines(
    text: str, avail_width: float, font: str, font_size: int, max_lines: int
) -> list[str]:
    """Wrap text into at most max_lines lines that each fit avail_width.

    If wrapping would produce more than max_lines, truncate to max_lines and
    append '…' to the last line, trimming characters from the right until the
    line + '…' fits. Each returned line is guaranteed to fit avail_width.
    """
    if not text:
        return []

    lines = simpleSplit(text, font, font_size, max(avail_width, 1))

    # simpleSplit can leave a CJK line wider than avail_width when there are
    # no break points it recognises — re-wrap those char-by-char.
    fixed: list[str] = []
    for line in lines:
        if stringWidth(line, font, font_size) <= avail_width:
            fixed.append(line)
            continue
        remaining = line
        while remaining:
            lo, hi = 1, len(remaining)
            while lo < hi:
                mid = (lo + hi + 1) // 2
                if stringWidth(remaining[:mid], font, font_size) <= avail_width:
                    lo = mid
                else:
                    hi = mid - 1
            fixed.append(remaining[:lo])
            remaining = remaining[lo:]

    if len(fixed) <= max_lines:
        return fixed

    # Truncation: concatenate the overflow tail back onto the last kept line,
    # then trim from the right until "<tail>…" fits.
    truncated = list(fixed[:max_lines])
    tail = truncated[-1] + "".join(fixed[max_lines:])
    ellipsis = "…"
    while tail and stringWidth(tail + ellipsis, font, font_size) > avail_width:
        tail = tail[:-1]
    truncated[-1] = (tail + ellipsis) if tail else ellipsis
    return truncated


def _format_chinese_date(dt) -> str:
    """Format like 2026年05月11日 — supplier-friendly, locale-stable."""
    if not dt:
        return ""
    return f"{dt.year}年{dt.month:02d}月{dt.day:02d}日"


def _rp_columns():
    """5-column layout for the 客户分拣 table.

    Returns (col_left_x, col_width) tuples — left edges and widths in pt."""
    widths = [_CONTENT_WIDTH * p for p in _RP_COL_PCT]
    lefts = [_MARGIN_X]
    for w in widths[:-1]:
        lefts.append(lefts[-1] + w)
    return lefts, widths


def _draw_handcraft_receipt_page(pdf, db, order) -> None:
    """Append the supplier-facing receipt page with a 5-column 客户分拣 table:
    商品名 / 数量 / 客户名 / 客户对应数量 / 拣货确认. 商品名 + 数量 span all
    customer rows of the same jewelry (rowspan); the right 3 cols have one
    row per customer. Real customer names are aliased to 客户 N."""
    from services.handcraft import get_handcraft_jewelry_breakdown

    groups = get_handcraft_jewelry_breakdown(db, order.id)
    payload_groups = []
    for g in groups:
        if g["kind"] != "jewelry":
            continue
        entries_with_customer = [e for e in g["entries"] if e["customer_name"]]
        if not entries_with_customer:
            continue
        payload_groups.append((g, entries_with_customer))

    if not payload_groups:
        return

    pdf.showPage()
    y = _PAGE_HEIGHT - _MARGIN_TOP

    # ── 1. Title: 手工回执 (large, bold, centered) ─────────────────
    pdf.setFont(_LABEL_FONT, 32)
    pdf.setFillColor(_RP_INK)
    tw = stringWidth(_RECEIPT_TITLE, _LABEL_FONT, 32)
    pdf.drawString((_PAGE_WIDTH - tw) / 2, y - 34, _RECEIPT_TITLE)
    y -= 52

    # Thin rule under title (full content width)
    pdf.setStrokeColor(_RP_RULE_SOFT)
    pdf.setLineWidth(0.8)
    pdf.line(_MARGIN_X, y, _PAGE_WIDTH - _MARGIN_X, y)
    y -= 26

    # ── 2. Receipt code block (HUGE, dominates the page identity) ──
    if order.receipt_code:
        # Small label above the code
        code_label = "回执编号"
        pdf.setFont(_LABEL_FONT, 11)
        pdf.setFillColor(_RP_INK_SOFT)
        lw = stringWidth(code_label, _LABEL_FONT, 11)
        pdf.drawString((_PAGE_WIDTH - lw) / 2, y - 12, code_label)
        y -= 22

        # The code itself — very large, makes the supplier impossible to mix
        # up. STSong-Light has no proper Latin advance, so without tracking
        # the chars visibly overlap. setCharSpace via TextObject preserves
        # the raw string for text extraction.
        #
        # CAREFUL: TextObject.setCharSpace mutates Canvas._charSpace as a
        # side-effect (see reportlab/pdfgen/textobject.py), so it leaks into
        # subsequent pdf.drawString() calls and tracks every char on the
        # rest of the page. We reset both ends: setCharSpace(0) inside the
        # text object so the BT/ET block ends clean, and pdf._charSpace = 0
        # after drawText as belt-and-suspenders.
        code_char_space = 10
        code_text_w = (
            stringWidth(order.receipt_code, _LABEL_FONT, 42)
            + code_char_space * (len(order.receipt_code) - 1)
        )
        code_obj = pdf.beginText((_PAGE_WIDTH - code_text_w) / 2, y - 36)
        code_obj.setFont(_LABEL_FONT, 42)
        code_obj.setFillColor(_RP_INK)
        code_obj.setCharSpace(code_char_space)
        code_obj.textOut(order.receipt_code)
        code_obj.setCharSpace(0)
        pdf.drawText(code_obj)
        pdf._charSpace = 0
        y -= 50

    # Thin rule under the code block
    pdf.setStrokeColor(_RP_RULE_SOFT)
    pdf.setLineWidth(0.8)
    pdf.line(_MARGIN_X, y, _PAGE_WIDTH - _MARGIN_X, y)
    y -= 22

    # ── 3. Meta strip: 手工商家 (left) / 发出日期 (right) ────────────
    pdf.setFont(_LABEL_FONT, 11)
    pdf.setFillColor(_RP_INK_SOFT)
    pdf.drawString(_MARGIN_X, y - 14, "手工商家:")
    pdf.setFillColor(_RP_INK)
    pdf.setFont(_LABEL_FONT, 12)
    pdf.drawString(_MARGIN_X + 56, y - 14, order.supplier_name or "")

    date_str = _format_chinese_date(order.created_at)
    pdf.setFont(_LABEL_FONT, 11)
    pdf.setFillColor(_RP_INK_SOFT)
    date_label = "发出日期:"
    pdf.setFont(_LABEL_FONT, 12)
    dw_val = stringWidth(date_str, _LABEL_FONT, 12)
    pdf.setFont(_LABEL_FONT, 11)
    dw_lab = stringWidth(date_label, _LABEL_FONT, 11)
    total_right_w = dw_lab + 8 + dw_val
    right_block_x = _PAGE_WIDTH - _MARGIN_X - total_right_w
    pdf.drawString(right_block_x, y - 14, date_label)
    pdf.setFont(_LABEL_FONT, 12)
    pdf.setFillColor(_RP_INK)
    pdf.drawString(right_block_x + dw_lab + 8, y - 14, date_str)
    y -= 26
    pdf.setStrokeColor(_RP_RULE_SOFT)
    pdf.setLineWidth(0.8)
    pdf.line(_MARGIN_X, y, _PAGE_WIDTH - _MARGIN_X, y)
    y -= 30

    # ── 4. Section heading: ── 客户分拣 ── (centered with side rules) ─
    heading = "客户分拣"
    pdf.setFont(_LABEL_FONT, 15)
    pdf.setFillColor(_RP_INK)
    hw = stringWidth(heading, _LABEL_FONT, 15)
    cx = _PAGE_WIDTH / 2
    pdf.drawString(cx - hw / 2, y - 16, heading)
    pdf.setStrokeColor(_RP_INK)
    pdf.setLineWidth(1.0)
    side_rule_w = 28
    side_gap = 14
    pdf.line(cx - hw / 2 - side_gap - side_rule_w, y - 10,
             cx - hw / 2 - side_gap, y - 10)
    pdf.line(cx + hw / 2 + side_gap, y - 10,
             cx + hw / 2 + side_gap + side_rule_w, y - 10)
    y -= 34

    # ── 5. Table ────────────────────────────────────────────────────
    _rp_draw_breakdown_table(pdf, payload_groups, y)

    # ── 6. Footer band at bottom of page ─────────────────────────────
    _rp_draw_footer_band(pdf)


def _rp_draw_breakdown_table(pdf, groups, y_start: float) -> float:
    """Render the 5-column table with rowspan-style merged 商品名 + 数量
    cells per jewelry group. Auto-paginates groups when they don't fit."""
    col_x, col_w = _rp_columns()
    right_edge = _PAGE_WIDTH - _MARGIN_X
    headers = ("商品名", "数量", "客户名", "客户对应数量", "拣货确认")

    def _draw_header(y: float) -> float:
        # Top border
        pdf.setStrokeColor(_RP_RULE)
        pdf.setLineWidth(0.7)
        pdf.line(_MARGIN_X, y, right_edge, y)
        # Header text — small caps-y feel, centered in each column
        pdf.setFont(_LABEL_FONT, 10)
        pdf.setFillColor(_RP_INK_SOFT)
        for i, label in enumerate(headers):
            tw = stringWidth(label, _LABEL_FONT, 10)
            pdf.drawString(col_x[i] + col_w[i] / 2 - tw / 2, y - 16, label)
        y -= _RP_HEADER_H
        pdf.setStrokeColor(_RP_RULE)
        pdf.setLineWidth(0.7)
        pdf.line(_MARGIN_X, y, right_edge, y)
        return y

    def _draw_verticals(top: float, bottom: float):
        pdf.setStrokeColor(_RP_RULE)
        pdf.setLineWidth(0.7)
        for i in range(5):
            pdf.line(col_x[i], top, col_x[i], bottom)
        pdf.line(right_edge, top, right_edge, bottom)

    # Reserve space at the bottom for the footer band (kept inside _MARGIN_BOTTOM)
    footer_reserve = 70
    table_top = y_start
    y = _draw_header(y_start)
    verticals_top = table_top  # verticals start at the very top border

    group_idx = 0
    for g, entries in groups:
        n = len(entries)
        group_h = n * _RP_ROW_H

        # Page break: close out this page's verticals, start fresh
        if y - group_h < _MARGIN_BOTTOM + footer_reserve:
            _draw_verticals(verticals_top, y)
            _rp_draw_footer_band(pdf)
            pdf.showPage()
            y_top = _PAGE_HEIGHT - _MARGIN_TOP
            y = _draw_header(y_top)
            verticals_top = y_top

        group_idx += 1
        group_top = y
        group_bottom = y - group_h
        cy = (group_top + group_bottom) / 2  # vertical center of the group

        # ── 商品名 cell: ① + jewelry name, vertically centered ──
        circle_r = 8
        cell0_left = col_x[0] + 12
        circle_cx = cell0_left + circle_r
        pdf.setStrokeColor(_RP_INK)
        pdf.setLineWidth(0.8)
        pdf.circle(circle_cx, cy + 1, circle_r, stroke=1, fill=0)
        pdf.setFont(_LABEL_FONT, 10)
        pdf.setFillColor(_RP_INK)
        idx_str = str(group_idx)
        iw = stringWidth(idx_str, _LABEL_FONT, 10)
        pdf.drawString(circle_cx - iw / 2, cy - 2, idx_str)

        name_x = circle_cx + circle_r + 8
        name_avail = (col_x[0] + col_w[0]) - name_x - 4
        name_lines = _wrap_text_to_lines(
            g["jewelry_name"] or "", name_avail, _LABEL_FONT, 12, max_lines=2
        )
        pdf.setFont(_LABEL_FONT, 12)
        if len(name_lines) <= 1:
            if name_lines:
                pdf.drawString(name_x, cy - 3, name_lines[0])
        else:
            line_gap = 14
            top_baseline = cy - 3 + (len(name_lines) - 1) * line_gap / 2
            for i, line in enumerate(name_lines):
                pdf.drawString(name_x, top_baseline - i * line_gap, line)

        # ── 数量 cell: total qty + 套, centered ──
        c1_cx = col_x[1] + col_w[1] / 2
        total_text = f"{_format_int(g['total_qty'])}套"
        pdf.setFont(_LABEL_FONT, 12)
        pdf.setFillColor(_RP_INK)
        tw_total = stringWidth(total_text, _LABEL_FONT, 12)
        pdf.drawString(c1_cx - tw_total / 2, cy - 3, total_text)

        # ── Customer rows: 客户名 / 客户对应数量 / 拣货确认 ──
        sorted_entries = sorted(entries, key=lambda x: x["hc_jewelry_item_id"])
        for cust_i, e in enumerate(sorted_entries, start=1):
            row_top = y
            row_bottom = y - _RP_ROW_H
            row_cy = (row_top + row_bottom) / 2

            # 客户名 — centered
            alias = f"客户 {cust_i}"
            pdf.setFont(_LABEL_FONT, 12)
            pdf.setFillColor(_RP_INK)
            aw = stringWidth(alias, _LABEL_FONT, 12)
            pdf.drawString(col_x[2] + col_w[2] / 2 - aw / 2, row_cy - 3, alias)

            # 客户对应数量 — centered
            qty_str = _format_int(e["qty"])
            qw = stringWidth(qty_str, _LABEL_FONT, 12)
            pdf.drawString(col_x[3] + col_w[3] / 2 - qw / 2, row_cy - 3, qty_str)

            # 拣货确认 — centered checkbox
            cb_size = 14
            chcx = col_x[4] + col_w[4] / 2
            pdf.setStrokeColor(_RP_INK)
            pdf.setLineWidth(0.9)
            pdf.rect(chcx - cb_size / 2, row_cy - cb_size / 2,
                     cb_size, cb_size, stroke=1, fill=0)

            # Inner row divider (only across right 3 cols — preserves rowspan
            # visual for 商品名 + 数量)
            if cust_i < n:
                pdf.setStrokeColor(_RP_RULE_SOFT)
                pdf.setLineWidth(0.4)
                pdf.line(col_x[2], row_bottom, right_edge, row_bottom)

            y -= _RP_ROW_H

        # Group-bottom full-width border
        pdf.setStrokeColor(_RP_RULE)
        pdf.setLineWidth(0.7)
        pdf.line(_MARGIN_X, y, right_edge, y)

    _draw_verticals(verticals_top, y)
    return y


def _rp_draw_footer_band(pdf) -> None:
    """Outlined reminder band at the very bottom of the page (no fill)."""
    band_h = 34
    band_bottom = _MARGIN_BOTTOM - 4
    pdf.setStrokeColor(_RP_FOOTER_RULE)
    pdf.setLineWidth(0.5)
    pdf.rect(_MARGIN_X, band_bottom, _CONTENT_WIDTH, band_h, stroke=1, fill=0)

    pdf.setFont(_LABEL_FONT, 11)
    pdf.setFillColor(_RP_INK_SOFT)
    nw = stringWidth(_RECEIPT_NOTE, _LABEL_FONT, 11)
    pdf.drawString((_PAGE_WIDTH - nw) / 2, band_bottom + (band_h - 11) / 2, _RECEIPT_NOTE)


def _format_int(n) -> str:
    """Format a numeric qty as integer when whole, else 2-decimal."""
    f = float(n)
    if f == int(f):
        return str(int(f))
    return f"{f:.2f}"
