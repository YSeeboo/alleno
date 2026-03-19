from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, UnidentifiedImageError
from sqlalchemy.orm import Session

from services.plating_export import (
    build_export_filename,
    download_image_bytes,
    format_excel_date,
    get_plating_export_payload,
)

_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "plating-order-template.xlsx"
_DETAIL_START_ROW = 8
_DETAIL_TEMPLATE_END_ROW = 12
_DELIVERY_SECTION_START_ROW = 13
_DEFAULT_ROW_HEIGHT = 15
_DEFAULT_COL_WIDTH = 8.43
_EMU_PER_PIXEL = 9525
_DETAIL_CELL_PADDING = 2
_DELIVERY_CELL_PADDING = 4
_DELIVERY_SLOTS = [
    {"start_col": 1, "end_col": 4, "label_row": 17, "image_row": 18, "end_row": 36},
    {"start_col": 5, "end_col": 6, "label_row": 17, "image_row": 18, "end_row": 36},
    {"start_col": 1, "end_col": 4, "label_row": 37, "image_row": 38, "end_row": 56},
    {"start_col": 5, "end_col": 6, "label_row": 37, "image_row": 38, "end_row": 56},
]


def build_plating_order_excel(db: Session, order_id: str) -> tuple[bytes, str]:
    if not _TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Plating Excel template not found: {_TEMPLATE_PATH}")
    payload = get_plating_export_payload(db, order_id)
    order = payload["order"]
    details = payload["details"]

    workbook = load_workbook(_TEMPLATE_PATH)
    sheet = workbook.active

    extra_rows = max(0, len(details) - (_DETAIL_TEMPLATE_END_ROW - _DETAIL_START_ROW + 1))
    if extra_rows:
        _expand_detail_rows(sheet, extra_rows)

    _fill_order_header(sheet, order.supplier_name, order.created_at)
    _fill_detail_rows(sheet, details)
    _fill_delivery_images(sheet, payload["delivery_images"], extra_rows)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue(), build_export_filename(order.supplier_name, order.created_at, "xlsx")


def _fill_order_header(sheet, supplier_name, created_at) -> None:
    supplier_name = (supplier_name or "").strip()
    sheet["A6"] = f"顾客: {supplier_name}" if supplier_name else "顾客: "
    sheet["C6"] = format_excel_date(created_at)


def _fill_detail_rows(sheet, details: list[dict]) -> None:
    for index, detail in enumerate(details):
        row = _DETAIL_START_ROW + index
        sheet.cell(row=row, column=1).value = detail["name"] or detail["part_id"]
        sheet.cell(row=row, column=3).value = detail["plating_method"] or None
        sheet.cell(row=row, column=4).value = detail["qty"]
        sheet.cell(row=row, column=5).value = detail["unit"] or None
        sheet.cell(row=row, column=6).value = detail["note"] or None
        _set_cell_font_size(sheet.cell(row=row, column=5), 18)
        _set_cell_font_size(sheet.cell(row=row, column=6), 18)
        if detail["part_image"]:
            _add_image(
                sheet,
                detail["part_image"],
                start_col=2,
                start_row=row,
                box_width=max(32, _column_pixel_width(sheet, 2)),
                box_height=max(32, _row_pixel_height(sheet, row)),
                padding=_DETAIL_CELL_PADDING,
            )


def _fill_delivery_images(sheet, delivery_images: list[str], row_shift: int) -> None:
    for index, source in enumerate(delivery_images[: len(_DELIVERY_SLOTS)]):
        slot = _DELIVERY_SLOTS[index]
        anchor_row = slot["image_row"] + row_shift
        start_col = slot["start_col"]
        _add_image(
            sheet,
            source,
            start_col=start_col,
            start_row=anchor_row,
            box_width=max(48, _columns_pixel_width(sheet, slot["start_col"], slot["end_col"])),
            box_height=max(
                48,
                _rows_pixel_height(sheet, slot["image_row"] + row_shift, slot["end_row"] + row_shift),
            ),
            padding=_DELIVERY_CELL_PADDING,
        )


def _expand_detail_rows(sheet, extra_rows: int) -> None:
    shifted_ranges = []
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= _DELIVERY_SECTION_START_ROW:
            shifted_ranges.append(
                (merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col)
            )
            sheet.unmerge_cells(str(merged_range))

    sheet.insert_rows(_DELIVERY_SECTION_START_ROW, extra_rows)

    for offset in range(extra_rows):
        _copy_row_format(sheet, _DETAIL_TEMPLATE_END_ROW, _DETAIL_TEMPLATE_END_ROW + 1 + offset)

    for min_row, min_col, max_row, max_col in shifted_ranges:
        sheet.merge_cells(
            start_row=min_row + extra_rows,
            start_column=min_col,
            end_row=max_row + extra_rows,
            end_column=max_col,
        )


def _copy_row_format(sheet, source_row: int, target_row: int) -> None:
    source_dimension = sheet.row_dimensions[source_row]
    target_dimension = sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for column in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(row=source_row, column=column)
        target_cell = sheet.cell(row=target_row, column=column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        target_cell.value = None


def _add_image(
    sheet,
    source: str,
    start_col: int,
    start_row: int,
    box_width: int,
    box_height: int,
    padding: int,
) -> None:
    image_bytes = download_image_bytes(source)
    if not image_bytes:
        return

    image = _build_excel_image(
        image_bytes,
        box_width=max(1, box_width - padding * 2),
        box_height=max(1, box_height - padding * 2),
    )
    if image is None:
        return

    offset_x = max(0, int((box_width - image.width) / 2))
    offset_y = max(0, int((box_height - image.height) / 2))
    image.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=start_col - 1,
            colOff=offset_x * _EMU_PER_PIXEL,
            row=start_row - 1,
            rowOff=offset_y * _EMU_PER_PIXEL,
        ),
        ext=XDRPositiveSize2D(image.width * _EMU_PER_PIXEL, image.height * _EMU_PER_PIXEL),
    )
    sheet.add_image(image)


def _build_excel_image(image_bytes: bytes, box_width: int, box_height: int) -> ExcelImage | None:
    try:
        with PILImage.open(BytesIO(image_bytes)) as raw_image:
            image = raw_image.copy()
    except (UnidentifiedImageError, OSError):
        return None

    width, height = image.size
    if width <= 0 or height <= 0:
        return None

    scale = min(box_width / width, box_height / height)
    target_width = max(1, int(width * scale))
    target_height = max(1, int(height * scale))
    if (target_width, target_height) != image.size:
        image = image.resize((target_width, target_height), PILImage.Resampling.LANCZOS)

    buffer = BytesIO()
    if image.mode not in ("RGB", "RGBA"):
        image = image.convert("RGBA")
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return ExcelImage(buffer)


def _column_pixel_width(sheet, column: int) -> int:
    width = sheet.column_dimensions[get_column_letter(column)].width
    width = _DEFAULT_COL_WIDTH if width is None else width
    return int(width * 7 + 5)


def _columns_pixel_width(sheet, start_col: int, end_col: int) -> int:
    return sum(_column_pixel_width(sheet, column) for column in range(start_col, end_col + 1))


def _row_pixel_height(sheet, row: int) -> int:
    height = sheet.row_dimensions[row].height
    height = _DEFAULT_ROW_HEIGHT if height is None else height
    return int(height * 4 / 3)


def _rows_pixel_height(sheet, start_row: int, end_row: int) -> int:
    return sum(_row_pixel_height(sheet, row) for row in range(start_row, end_row + 1))


def _set_cell_font_size(cell, size: int) -> None:
    font = copy(cell.font)
    font.sz = size
    cell.font = font
